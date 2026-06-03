# -*- coding: utf-8 -*-
"""
DocParser Pro - 免费云部署版 (简化版)
适配 Vercel - 使用阿里云文档智能API
"""

import os
import sys
import uuid
import json
import base64
import tempfile
from datetime import datetime
from pathlib import Path

from flask import Flask, render_template, request, jsonify, Response

app = Flask(__name__, template_folder='../templates', static_folder='../static')

# 阿里云配置
ALIBABA_ACCESS_KEY = 'LTAI5t…H42N'
ALIBABA_SECRET = 'YBT7xj…0Y1c'

# 定价
PRICE_PER_PAGE = 0.25

# 临时存储
temp_storage = {}

@app.route('/')
def index():
    """首页"""
    return render_template('index.html')

@app.route('/api/health')
def health():
    """健康检查"""
    return jsonify({
        'status': 'ok',
        'time': datetime.now().isoformat(),
        'alibaba_configured': bool(ALIBABA_ACCESS_KEY and ALIBABA_SECRET)
    })

@app.route('/api/upload', methods=['POST'])
def api_upload():
    """上传PDF"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '没有选择文件'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': '文件名为空'})
    
    if not file.filename.lower().endswith('.pdf'):
        return jsonify({'success': False, 'message': '只支持PDF文件'})
    
    output_format = request.form.get('format', 'excel')
    task_id = str(uuid.uuid4())[:8]
    
    try:
        # 读取PDF
        pdf_content = file.read()
        
        # 获取页数
        try:
            import PyPDF2
            from io import BytesIO
            pdf_reader = PyPDF2.PdfReader(BytesIO(pdf_content))
            page_count = len(pdf_reader.pages)
        except:
            page_count = 1
        
        cost = page_count * PRICE_PER_PAGE
        
        temp_storage[task_id] = {
            'pdf_content': base64.b64encode(pdf_content).decode(),
            'filename': file.filename,
            'page_count': page_count,
            'format': output_format,
            'cost': cost,
            'status': 'pending'
        }
        
        return jsonify({
            'success': True,
            'task_id': task_id,
            'message': f'文件上传成功！共{page_count}页，费用{cost:.2f}元',
            'pages': page_count,
            'cost': cost
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'上传失败: {str(e)}'})

def call_aliyun_ocr(pdf_path):
    """调用阿里云OCR"""
    try:
        # 使用阿里云官方SDK
        from alibabacloud_ocr_api20210707.client import Client as OcrClient
        from alibabacloud_tea_openapi import models as open_api_models
        
        config = open_api_models.Config(
            access_key_id=ALIBABA_ACCESS_KEY,
            access_key_secret=ALIBABA_SECRET
        )
        config.endpoint = 'ocr-api.cn-hangzhou.aliyuncs.com'
        client = OcrClient(config)
        
        # 读取PDF并转为base64
        with open(pdf_path, 'rb') as f:
            pdf_base64 = base64.b64encode(f.read()).decode()
        
        # 调用OCR
        from alibabacloud_ocr_api20210707 import models as ocr_models
        request = ocr_models.RecognizeAllTextRequest(
            body=ocr_models.RecognizeAllTextRequestBody(
                url=pdf_base64,
                type='pdf'
            )
        )
        response = client.recognize_all_text(request)
        return response.body
        
    except Exception as e:
        print(f"OCR调用失败: {e}")
        return None

@app.route('/api/convert/<task_id>', methods=['POST'])
def api_convert(task_id):
    """执行转换"""
    if task_id not in temp_storage:
        return jsonify({'success': False, 'message': '任务不存在'})
    
    task = temp_storage[task_id]
    
    if not ALIBABA_ACCESS_KEY or not ALIBABA_SECRET:
        return jsonify({
            'success': False, 
            'message': 'OCR服务未配置，请联系管理员'
        })
    
    try:
        # 解码PDF
        pdf_content = base64.b64decode(task['pdf_content'])
        
        # 保存临时文件
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
            tmp.write(pdf_content)
            tmp_path = tmp.name
        
        # 调用OCR
        ocr_result = call_aliyun_ocr(tmp_path)
        
        if not ocr_result:
            return jsonify({'success': False, 'message': 'OCR识别失败'})
        
        # 生成简单的Excel（实际应该解析OCR结果）
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "识别结果"
        
        # 添加识别结果
        ws['A1'] = "PDF识别结果"
        ws['A2'] = f"文件名: {task['filename']}"
        ws['A3'] = f"页数: {task['page_count']}"
        ws['A4'] = f"识别时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A6'] = "OCR原始结果:"
        ws['A7'] = str(ocr_result)[:1000]  # 简化显示
        
        # 保存Excel
        output_path = tmp_path.replace('.pdf', '.xlsx')
        wb.save(output_path)
        
        # 读取结果
        with open(output_path, 'rb') as f:
            output_content = f.read()
        
        task['status'] = 'completed'
        task['output_content'] = base64.b64encode(output_content).decode()
        task['output_filename'] = task['filename'].replace('.pdf', '.xlsx')
        
        # 清理
        os.unlink(tmp_path)
        os.unlink(output_path)
        
        return jsonify({
            'success': True,
            'message': '转换成功！',
            'download_url': f'/api/download/{task_id}'
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        task['status'] = 'failed'
        return jsonify({'success': False, 'message': f'转换失败: {str(e)}'})

@app.route('/api/download/<task_id>')
def api_download(task_id):
    """下载文件"""
    if task_id not in temp_storage:
        return jsonify({'success': False, 'message': '文件不存在'}), 404
    
    task = temp_storage[task_id]
    
    if task.get('status') != 'completed':
        return jsonify({'success': False, 'message': '转换未完成'}), 400
    
    output_content = base64.b64decode(task['output_content'])
    output_filename = task.get('output_filename', 'converted.xlsx')
    
    return Response(
        output_content,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={output_filename}'}
    )

@app.route('/api/status/<task_id>')
def api_status(task_id):
    """查询状态"""
    if task_id not in temp_storage:
        return jsonify({'success': False, 'message': '任务不存在'})
    
    task = temp_storage[task_id]
    return jsonify({
        'success': True,
        'status': task.get('status', 'unknown')
    })

app.debug = False

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
